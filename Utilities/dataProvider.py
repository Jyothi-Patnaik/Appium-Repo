import openpyxl


def get_data(excel):
    workbook = openpyxl.load_workbook("C://Users//admin//PycharmProjects//Inlynk_Appium//Excel//Book1.xlsx")
    sheet = workbook[excel]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    print("total cols are ",str(totalcols))
    print("total rows are ", str(totalrows))
    mainList = []

    for i in range(2,totalrows+1):
        dataList = []
        for j in range(1, totalcols+1):
            data = sheet.cell(row=i,column=j).value
            dataList.insert(j,data)
        mainList.insert(i,dataList)
        print(mainList)
    return mainList















































































